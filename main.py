"""
Excel Filler Service — Railway deployment
Accepts an Excel template (base64) + target_year,
fills it from Supabase indicator_values, and returns
the filled Excel uploaded to Supabase Storage.
"""
import os, re, json, base64, io, requests
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string

app = FastAPI(title="Excel Filler Service")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Supabase config (from env vars) ──────────────────────────────────────────
SB_URL  = os.environ["SUPABASE_URL"]
SB_KEY  = os.environ["SUPABASE_KEY"]
SB_HDRS = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}
SB_UPSERT_HDRS = {**SB_HDRS, "Content-Type": "application/json",
                  "Prefer": "resolution=merge-duplicates,return=minimal"}

# ── Sheet layouts ─────────────────────────────────────────────────────────────
LAYOUTS = {
    'Kuwait':  {'cc': 'KW',  'col2026': 'K', 'col2027': 'M', 'oilStart': 15},
    'KSA':     {'cc': 'SA',  'col2026': 'K', 'col2027': 'M', 'oilStart': 15},
    'UAE':     {'cc': 'AE',  'col2026': 'I', 'col2027': 'K', 'oilStart': 13},
    'Oman':    {'cc': 'OM',  'col2026': 'I', 'col2027': 'K', 'oilStart': 13},
    'Bahrain': {'cc': 'BH',  'col2026': 'I', 'col2027': 'K', 'oilStart': 13},
    'Qatar':   {'cc': 'QA',  'col2026': 'H', 'col2027': 'J', 'oilStart': 12},
    'GCC':     {'cc': 'GCC', 'col2026': 'J', 'col2027': 'L', 'oilStart': 14},
}

EXCEL_INDUSTRIES = [
    'Oil, Gas and Energy', 'Financial Services', 'F&B', 'Healthcare',
    'Telecom and Technology', 'Real Estate', 'Insurance', 'Retail and Consumer Goods',
    'Manufacturing and Industrial', 'Consulting', 'Construction',
    'Tourism and Hospitality', 'Education',
]

INDUSTRY_MAP = {
    'Oil and Gas': 'Oil, Gas and Energy', 'Oil, Gas, and Mining': 'Oil, Gas and Energy',
    'Utilities': 'Oil, Gas and Energy', 'Electric Power Generation': 'Oil, Gas and Energy',
    'Banking': 'Financial Services', 'Investment Banking': 'Financial Services',
    'Capital Markets': 'Financial Services', 'Insurance Carriers': 'Financial Services',
    'Financial Services': 'Financial Services', 'Credit Intermediation': 'Financial Services',
    'Investment Management': 'Financial Services',
    'Food and Beverage Services': 'F&B', 'Restaurants': 'F&B',
    'Food and Beverage Manufacturing': 'F&B', 'Food and Beverage Retail': 'F&B',
    'Hospitals and Health Care': 'Healthcare', 'Hospitals': 'Healthcare',
    'Medical Practices': 'Healthcare', 'Medical Equipment Manufacturing': 'Healthcare',
    'Pharmaceutical Manufacturing': 'Healthcare', 'Public Health': 'Healthcare',
    'IT Services and IT Consulting': 'Telecom and Technology',
    'Software Development': 'Telecom and Technology',
    'Telecommunications': 'Telecom and Technology',
    'Technology Information and Media': 'Telecom and Technology',
    'Computer Hardware Manufacturing': 'Telecom and Technology',
    'Computer and Network Security': 'Telecom and Technology',
    'Real Estate': 'Real Estate', 'Insurance': 'Insurance',
    'Retail': 'Retail and Consumer Goods', 'Retail Apparel and Fashion': 'Retail and Consumer Goods',
    'Consumer Services': 'Retail and Consumer Goods', 'Wholesale': 'Retail and Consumer Goods',
    'Manufacturing': 'Manufacturing and Industrial',
    'Industrial Machinery Manufacturing': 'Manufacturing and Industrial',
    'Chemical Manufacturing': 'Manufacturing and Industrial',
    'Motor Vehicle Manufacturing': 'Manufacturing and Industrial',
    'Business Consulting and Services': 'Consulting', 'Professional Services': 'Consulting',
    'Staffing and Recruiting': 'Consulting', 'Human Resources Services': 'Consulting',
    'Construction': 'Construction', 'Civil Engineering': 'Construction',
    'Specialty Trade Contractors': 'Construction',
    'Hospitality': 'Tourism and Hospitality', 'Hotels and Motels': 'Tourism and Hospitality',
    'Travel Arrangements': 'Tourism and Hospitality', 'Airlines and Aviation': 'Tourism and Hospitality',
    'Education': 'Education', 'Higher Education': 'Education',
    'Primary and Secondary Education': 'Education', 'E-Learning Providers': 'Education',
}

DEMAND_PRIORITY = {'Very high': 4, 'High': 3, 'Moderate': 2, 'Low': 1}

NAME_OVERRIDES = {
    "average net profit (usd) thousands": "bod__average_net_profit_in_usd_thousands",
    "average total remuneration (usd) thousands": "bod__average_total_remuneration_in_usd_thousands",
    "job seekers' gender": "talent_mobility__job_seekers_gender",
    "job seekers' generation": "talent_mobility__job_seekers_generation",
    "job seekers' nationality": "talent_mobility__job_seekers_nationality",
    "job seekers' qualification": "talent_mobility__job_seekers_qualification",
    "job seekers' job level": "talent_mobility__job_seekers_job_level",
    "organizations providing long term incentive plans (ltips)": "rewards__organizations_providing_long_term_incentive_plans_ltips",
    "organizations providing non-monetary benefits": "rewards__organizations_providing_non_monetary_benefits",
    "retention strategies implementation": "talent_mobility__retention_strategy_implementation",
    "succession plan rating": "talent_mobility__succession_planning_readiness",
    "talent work-life balance": "talent_mobility__talent_work_life_balance",
    "time to hire (days)": "talent_mobility__time_to_hire_days",
    "types of non-monetary benefits": "rewards__types_of_non_monetary_benefits",
    "organizations facing skill shortages": "talent_mobility__organizations_facing_skill_shortages",
    "organizations adopting remote work policies": "talent_mobility__organizations_adopting_remote_work_policies",
    "organizations providing annual bonus": "rewards__organizations_providing_annual_bonus",
    "organizations providing annual increments": "rewards__organizations_providing_annual_increments",
    "organizations planning to provide ltips": "rewards__organizations_planning_to_provide_ltips",
    "organizations planning to provide annual bonus": "rewards__organizations_planning_to_provide_annual_bonus",
    "organizations planning to  provide annual bonus": "rewards__organizations_planning_to_provide_annual_bonus",
    "percentage of increments provided to employees": "rewards__percentage_of_increments_provided_to_employees",
    "talent salary negotiation attempts": "rewards__talent_salary_negotiation_attempts",
    "employee training programs": "esg__average_hours_of_training_provided_to_employees",
}

NESTED_INDICATOR_NAMES = {
    "job seekers' nationality": "talent_mobility__job_seekers_nationality",
    "average employee tenure": "talent_mobility__average_employee_tenure",
    "talent priorities before job acceptance": "talent_mobility__talent_priorities_before_job_acceptance",
    "organizations providing promotions to ftes": "talent_mobility__organizations_providing_promotions_to_ftes",
    "talent salary negotiation attempts": "rewards__talent_salary_negotiation_attempts",
    "inflation": "economic__inflation",
}

HEADER_ONLY_KEYS = {
    'rewards__organizations_providing_annual_bonus',
    'rewards__organizations_providing_annual_increments',
    'rewards__organizations_providing_long_term_incentive_plans_ltips',
    'rewards__organizations_planning_to_provide_ltips',
    'rewards__percentage_of_increments_provided_to_employees',
    'rewards__organizations_providing_non_monetary_benefits',
    'talent_mobility__organizations_facing_skill_shortages',
    'talent_mobility__organizations_adopting_remote_work_policies',
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def normalize(s):
    if not s: return ''
    s = str(s).strip().lower()
    s = re.sub(r"['\u2018\u2019\u02bc\u0060\u00b4]", "'", s)
    s = re.sub(r'[\u2010\u2011\u2012\u2013\u2014\u2015\u2212]', '-', s)
    s = re.sub(r'\s+', ' ', s)
    return s

def norm_sub(s):
    if s is None: return ''
    return str(s).strip().title()

def col_idx(letter):
    return column_index_from_string(letter)

def write_val(ws, row, col_letter, value):
    try:
        cell = ws.cell(row, col_idx(col_letter))
        if isinstance(cell, MergedCell): return False
        if value is None: return False
        cell.value = value
        return True
    except Exception:
        return False

def fetch_all(endpoint_path):
    rows, offset = [], 0
    while True:
        r = requests.get(
            f"{SB_URL}/rest/v1/{endpoint_path}&limit=1000&offset={offset}",
            headers=SB_HDRS
        )
        batch = r.json()
        if not isinstance(batch, list) or not batch: break
        rows.extend(batch)
        if len(batch) < 1000: break
        offset += 1000
    return rows

# ── Request model ─────────────────────────────────────────────────────────────
class FillRequest(BaseModel):
    excel_base64: str
    excel_filename: str = "output.xlsx"
    target_year: int = 2026
    bucket: str = "pipeline-uploads"
    run_id: str = ""

# ── Main endpoint ─────────────────────────────────────────────────────────────
@app.post("/fill")
def fill_excel(req: FillRequest):
    target_year = req.target_year
    forecast_year_str = str(target_year)
    prev_year_str = str(target_year - 1)

    # 1. Decode Excel
    try:
        excel_bytes = base64.b64decode(req.excel_base64)
    except Exception as e:
        raise HTTPException(400, f"Invalid base64: {e}")

    # 2. Fetch data from Supabase
    forecast_rows = fetch_all(
        f"indicator_values?value_type=eq.forecast"
        f"&select=indicator_key,country_code,industry,sub_row,year,value"
    )
    measured_rows = fetch_all(
        f"indicator_values?value_type=eq.measured"
        f"&select=indicator_key,country_code,industry,sub_row,year,value"
    )

    # 3. Fetch indicator catalog
    cat_rows = fetch_all("indicator_catalog?select=indicator_key,name")
    NAME_TO_KEY = {c['name'].strip().lower(): c['indicator_key'] for c in cat_rows}
    NORM_TO_KEY = {normalize(c['name']): c['indicator_key'] for c in cat_rows}

    def lookup_key(name):
        lo = name.strip().lower()
        lo_norm = re.sub(r"['''\u02bc]", "'", lo)
        if lo_norm in NAME_OVERRIDES: return NAME_OVERRIDES[lo_norm]
        if lo in NAME_TO_KEY: return NAME_TO_KEY[lo]
        norm = normalize(name)
        if norm in NORM_TO_KEY: return NORM_TO_KEY[norm]
        for cat_norm, key in NORM_TO_KEY.items():
            if norm in cat_norm or cat_norm in norm: return key
        return None

    # 4. Build FV index (target_year values)
    FV = {}
    def fv_key(row):
        return f"{row['indicator_key']}|{row['country_code']}|{norm_sub(row['sub_row'])}|{row['year']}"

    for row in forecast_rows:
        k = fv_key(row)
        existing = FV.get(k)
        if existing is None:
            FV[k] = row['value']
        elif row['industry'] in ('All', None, ''):
            FV[k] = row['value']

    for row in measured_rows:
        if row['year'] == target_year:
            k = fv_key(row)
            existing = FV.get(k)
            if existing is None:
                FV[k] = row['value']
            elif row['industry'] in ('All', None, ''):
                FV[k] = row['value']

    # 5. Build LATEST index (fallback for non-forecasted indicators)
    LATEST = {}
    for row in measured_rows:
        k = f"{row['indicator_key']}|{row['country_code']}|{norm_sub(row['sub_row'])}"
        existing = LATEST.get(k)
        if existing is None or row['year'] > existing[0]:
            if existing is None or row['industry'] in ('All', None, '') or existing[0] < row['year']:
                LATEST[k] = (row['year'], row['value'])

    # 6. Build industry-level FV index
    FV_IND = {}
    for row in forecast_rows + [r for r in measured_rows if r['year'] == target_year]:
        if row['industry'] and row['industry'] not in ('All', ''):
            k = f"{row['indicator_key']}|{row['country_code']}|{row['industry']}|{norm_sub(row['sub_row'])}|{row['year']}"
            if k not in FV_IND:
                FV_IND[k] = row['value']

    # 6b. Build LATEST_IND index (fallback for per-industry when no forecast exists)
    LATEST_IND = {}
    for row in measured_rows:
        if row['industry'] and row['industry'] not in ('All', ''):
            k = f"{row['indicator_key']}|{row['country_code']}|{row['industry']}|{norm_sub(row['sub_row'])}"
            existing = LATEST_IND.get(k)
            if existing is None or row['year'] > existing[0]:
                LATEST_IND[k] = (row['year'], row['value'])

    FORECAST_KEYS = set(r['indicator_key'] for r in forecast_rows)

    # 7. Fetch demand data from Supabase
    demand_rows = fetch_all(
        "indicator_values?indicator_key=eq.talent_mobility__industries_in_high_or_lower_demand"
        "&select=country_code,sub_row,value"
    )
    CC_TO_SHEET = {'KW':'Kuwait','SA':'KSA','AE':'UAE','OM':'Oman','BH':'Bahrain','QA':'Qatar','GCC':'GCC'}
    demand_by_country = {s: {} for s in LAYOUTS}
    for row in demand_rows:
        sheet = CC_TO_SHEET.get(row.get('country_code', ''))
        if not sheet: continue
        excel_ind = INDUSTRY_MAP.get(row.get('sub_row', ''))
        if not excel_ind: continue
        demand = row.get('value', '')
        cur = demand_by_country[sheet].get(excel_ind, '')
        if DEMAND_PRIORITY.get(str(demand), 0) > DEMAND_PRIORITY.get(cur, 0):
            demand_by_country[sheet][excel_ind] = demand

    # 8. Open workbook and fill
    wb = load_workbook(io.BytesIO(excel_bytes))
    total_written = 0
    unmatched = set()

    for sheet_name, layout in LAYOUTS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cc = layout['cc']
        col26 = layout['col2026']
        col27 = layout['col2027']
        oil_start = layout['oilStart']
        written = 0
        current_key = None
        is_forecasted = False

        for row in range(4, 350):
            c_cell = ws.cell(row, 3)
            d_cell = ws.cell(row, 4)
            if isinstance(d_cell, MergedCell): continue
            d_val = d_cell.value
            c_val = c_cell.value if not isinstance(c_cell, MergedCell) else None
            if d_val is None: continue
            d_str = str(d_val).strip()

            # Indicator header row
            if c_val is not None and str(c_val).strip().isdigit():
                current_key = lookup_key(d_str)
                if not current_key:
                    unmatched.add(d_str)
                is_forecasted = current_key in FORECAST_KEYS if current_key else False

                # Industries demand row
                if current_key and 'industries_in_high_or_lower_demand' in current_key:
                    demand = demand_by_country.get(sheet_name, {})
                    for i, excel_ind in enumerate(EXCEL_INDUSTRIES):
                        label = demand.get(excel_ind)
                        if not label: continue
                        col_ind = get_column_letter(oil_start + (i * 4) + 2)
                        if write_val(ws, row, col_ind, label): written += 1
                    continue

                # Header-only indicators
                if current_key and current_key in HEADER_ONLY_KEYS:
                    k26 = f"{current_key}|{cc}|Overall|{target_year}"
                    val26 = FV.get(k26)
                    if val26 is None:
                        le = LATEST.get(f"{current_key}|{cc}|Overall")
                        if le and le[0] == target_year: val26 = le[1]  # no older-year carry-forward
                    k27 = f"{current_key}|{cc}|Overall|{target_year + 1}"
                    val27 = FV.get(k27)
                    if val26 is not None:
                        if write_val(ws, row, col26, val26): written += 1
                    if val27 is not None:
                        if write_val(ws, row, col27, val27): written += 1
                    for i, excel_ind in enumerate(EXCEL_INDUSTRIES):
                        col_ind26 = get_column_letter(oil_start + (i * 4) + 2)
                        col_ind27 = get_column_letter(oil_start + (i * 4) + 3)
                        v26 = FV_IND.get(f"{current_key}|{cc}|{excel_ind}|Overall|{target_year}")
                        if v26 is None:
                            li = LATEST_IND.get(f"{current_key}|{cc}|{excel_ind}|Overall")
                            if li and li[0] == target_year: v26 = li[1]  # no older-year carry-forward
                        v27 = FV_IND.get(f"{current_key}|{cc}|{excel_ind}|Overall|{target_year + 1}")
                        if v26 is not None:
                            if write_val(ws, row, col_ind26, v26): written += 1
                        if v27 is not None:
                            if write_val(ws, row, col_ind27, v27): written += 1
                    continue
                continue

            # Sub-row data row
            if not current_key: continue
            sub_row = d_str

            # Check for nested indicator header
            sub_lo = re.sub(r"['''\u02bc]", "'", sub_row.strip().lower())
            if sub_lo in NESTED_INDICATOR_NAMES:
                current_key = NESTED_INDICATOR_NAMES[sub_lo]
                is_forecasted = current_key in FORECAST_KEYS
                if current_key in HEADER_ONLY_KEYS:
                    k26 = f"{current_key}|{cc}|Overall|{target_year}"
                    val26 = FV.get(k26)
                    if val26 is None:
                        le = LATEST.get(f"{current_key}|{cc}|Overall")
                        if le and le[0] == target_year: val26 = le[1]  # no older-year carry-forward
                    val27 = FV.get(f"{current_key}|{cc}|Overall|{target_year + 1}")
                    if val26 is not None:
                        if write_val(ws, row, col26, val26): written += 1
                    if val27 is not None:
                        if write_val(ws, row, col27, val27): written += 1
                continue

            sub_norm = norm_sub(sub_row)
            k26 = f"{current_key}|{cc}|{sub_norm}|{target_year}"
            k27 = f"{current_key}|{cc}|{sub_norm}|{target_year + 1}"

            val26 = FV.get(k26)
            if val26 is None:
                le = LATEST.get(f"{current_key}|{cc}|{sub_norm}")
                if le and le[0] == target_year: val26 = le[1]  # no older-year carry-forward
            val27 = FV.get(k27)

            if val26 is not None:
                if write_val(ws, row, col26, val26): written += 1
            if val27 is not None:
                if write_val(ws, row, col27, val27): written += 1

            for i, excel_ind in enumerate(EXCEL_INDUSTRIES):
                col_ind26 = get_column_letter(oil_start + (i * 4) + 2)
                col_ind27 = get_column_letter(oil_start + (i * 4) + 3)
                v26 = FV_IND.get(f"{current_key}|{cc}|{excel_ind}|{sub_norm}|{target_year}")
                if v26 is None:
                    li = LATEST_IND.get(f"{current_key}|{cc}|{excel_ind}|{sub_norm}")
                    if li and li[0] == target_year: v26 = li[1]  # no older-year carry-forward
                v27 = FV_IND.get(f"{current_key}|{cc}|{excel_ind}|{sub_norm}|{target_year + 1}")
                if v26 is not None:
                    if write_val(ws, row, col_ind26, v26): written += 1
                if v27 is not None:
                    if write_val(ws, row, col_ind27, v27): written += 1

        total_written += written

    # 9. Save filled workbook to bytes
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    filled_bytes = out_buf.read()
    filled_b64 = base64.b64encode(filled_bytes).decode()

    # 10. Upload to Supabase Storage
    run_id = req.run_id or f"run-{target_year}"
    filename = f"filled-{run_id}-{req.excel_filename}"
    bucket = req.bucket

    upload_resp = requests.post(
        f"{SB_URL}/storage/v1/object/{bucket}/{filename}",
        headers={
            **SB_HDRS,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "x-upsert": "true"
        },
        data=filled_bytes
    )

    if upload_resp.status_code not in (200, 201):
        # Return base64 as fallback if storage upload fails
        return {
            "ok": True,
            "total_written": total_written,
            "unmatched_count": len(unmatched),
            "download_url": None,
            "excel_base64": filled_b64,
            "filename": filename,
            "storage_error": upload_resp.text
        }

    # Generate signed URL (valid 7 days)
    sign_resp = requests.post(
        f"{SB_URL}/storage/v1/object/sign/{bucket}/{filename}",
        headers={**SB_HDRS, "Content-Type": "application/json"},
        json={"expiresIn": 604800}
    )
    download_url = None
    if sign_resp.status_code == 200:
        token = sign_resp.json().get("signedURL") or sign_resp.json().get("signedUrl", "")
        download_url = f"{SB_URL}/storage/v1{token}" if token.startswith("/") else token

    return {
        "ok": True,
        "total_written": total_written,
        "unmatched_count": len(unmatched),
        "download_url": download_url,
        "filename": filename,
        "target_year": target_year
    }

@app.get("/health")
def health():
    return {"status": "ok"}
